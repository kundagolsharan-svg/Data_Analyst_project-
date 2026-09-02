"""
03_build_excel_dashboard.py
-----------------------------
Data Analyst Task 1: Excel / Google Sheets.

Builds a multi-sheet sales dashboard workbook that demonstrates:
  - Data Cleaning (a documented "Raw Data" table, already cleaned)
  - Sorting & Filtering (AutoFilter enabled on the data table)
  - Conditional Formatting (color scales + data bars + highlight rules)
  - Pivot-style summary tables built with SUMIFS/AVERAGEIFS/COUNTIFS
    (formula-driven, so they recalculate — a true Excel PivotTable object
    can't be authored by openpyxl, so this mirrors one with live formulas
    the user can also turn into a real PivotTable via Insert > PivotTable)
  - Charts (bar, line, pie)
  - Lookup Functions (INDEX/MATCH — XLOOKUP is intentionally avoided; see
    note in the Lookup Demo sheet, since it isn't supported by every Excel
    version/verifier)
  - A Dashboard sheet with KPI cards pulling live formulas

Run: python3 03_build_excel_dashboard.py
Output: ../excel/sales_dashboard.xlsx
"""

from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart.label import DataLabelList

BASE_DIR = Path(__file__).resolve().parent.parent
CLEAN_PATH = BASE_DIR / "data" / "sales_data_cleaned.csv"
OUT_PATH = BASE_DIR / "excel" / "sales_dashboard.xlsx"
(BASE_DIR / "excel").mkdir(parents=True, exist_ok=True)

df = pd.read_csv(CLEAN_PATH, parse_dates=["OrderDate", "ShipDate"])
df = df.sort_values("OrderDate").reset_index(drop=True)

# Trim to the columns most useful for a business dashboard
cols = ["OrderID", "OrderDate", "CustomerName", "Region", "Category", "Product",
        "ShipMode", "Quantity", "UnitPrice", "Discount", "Sales", "Profit", "ProfitMargin"]
data = df[cols].copy()

wb = Workbook()

# ---------------------------------------------------------------------------
# Shared styles
# ---------------------------------------------------------------------------
FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color="1F4E78")
SUBTITLE_FONT = Font(name=FONT_NAME, italic=True, size=10, color="666666")
KPI_LABEL_FONT = Font(name=FONT_NAME, size=10, color="666666")
KPI_VALUE_FONT = Font(name=FONT_NAME, bold=True, size=18, color="1F4E78")
SECTION_FONT = Font(name=FONT_NAME, bold=True, size=12, color="1F4E78")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
KPI_FILL = PatternFill("solid", fgColor="EFF6FF")

def style_header_row(ws, row, first_col, last_col):
    for c in range(first_col, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

# ---------------------------------------------------------------------------
# Sheet 1: Raw Data (cleaned) — sorting/filtering + conditional formatting
# ---------------------------------------------------------------------------
ws_data = wb.active
ws_data.title = "Raw Data"

ws_data.append(list(data.columns))
for row in data.itertuples(index=False):
    row = list(row)
    row[1] = row[1].strftime("%Y-%m-%d") if pd.notna(row[1]) else None  # OrderDate as string-safe write
    ws_data.append(row)

n_rows = len(data) + 1
n_cols = len(data.columns)
style_header_row(ws_data, 1, 1, n_cols)

# proper date type + currency formats
date_col = data.columns.get_loc("OrderDate") + 1
for r in range(2, n_rows + 1):
    ws_data.cell(row=r, column=date_col).number_format = "yyyy-mm-dd"
for col_name in ["UnitPrice", "Sales", "Profit"]:
    c = data.columns.get_loc(col_name) + 1
    for r in range(2, n_rows + 1):
        ws_data.cell(row=r, column=c).number_format = "$#,##0.00"
disc_col = data.columns.get_loc("Discount") + 1
margin_col = data.columns.get_loc("ProfitMargin") + 1
for r in range(2, n_rows + 1):
    ws_data.cell(row=r, column=disc_col).number_format = "0.0%"
    ws_data.cell(row=r, column=margin_col).number_format = "0.0%"

# AutoFilter (sorting & filtering) + freeze header
ws_data.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows}"
ws_data.freeze_panes = "A2"

# Excel Table styling
tbl = Table(displayName="SalesData", ref=f"A1:{get_column_letter(n_cols)}{n_rows}")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws_data.add_table(tbl)

# Conditional formatting: color scale on Profit, data bars on Sales
profit_col_letter = get_column_letter(data.columns.get_loc("Profit") + 1)
sales_col_letter = get_column_letter(data.columns.get_loc("Sales") + 1)
ws_data.conditional_formatting.add(
    f"{profit_col_letter}2:{profit_col_letter}{n_rows}",
    ColorScaleRule(start_type="min", start_color="F8696B",
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",
                    end_type="max", end_color="63BE7B"),
)
ws_data.conditional_formatting.add(
    f"{sales_col_letter}2:{sales_col_letter}{n_rows}",
    DataBarRule(start_type="min", end_type="max", color="638EC6"),
)
# Highlight rows where Discount is high (>=20%)
disc_col_letter = get_column_letter(disc_col)
ws_data.conditional_formatting.add(
    f"{disc_col_letter}2:{disc_col_letter}{n_rows}",
    CellIsRule(operator="greaterThanOrEqual", formula=["0.2"],
               fill=PatternFill("solid", fgColor="FFF2CC")),
)

for c in range(1, n_cols + 1):
    max_len = max(len(str(data.columns[c - 1])), 10)
    ws_data.column_dimensions[get_column_letter(c)].width = max_len + 4

# ---------------------------------------------------------------------------
# Sheet 2: Summary (pivot-style tables, formula-driven with SUMIFS/AVERAGEIFS)
# ---------------------------------------------------------------------------
ws_sum = wb.create_sheet("Summary")
ws_sum["A1"] = "Pivot-Style Summary Tables"
ws_sum["A1"].font = TITLE_FONT
ws_sum["A2"] = "All values below are live SUMIFS/AVERAGEIFS/COUNTIFS formulas referencing 'Raw Data'. Select this range and Insert > PivotTable for a native pivot if preferred."
ws_sum["A2"].font = SUBTITLE_FONT
ws_sum.merge_cells("A2:F2")

regions = sorted(data["Region"].dropna().unique().tolist())
categories = sorted(data["Category"].dropna().unique().tolist())
months = sorted(data["OrderDate"].dt.to_period("M").astype(str).unique().tolist())

data_range = f"'Raw Data'!"
region_rng = f"{data_range}${get_column_letter(data.columns.get_loc('Region')+1)}$2:${get_column_letter(data.columns.get_loc('Region')+1)}${n_rows}"
category_rng = f"{data_range}${get_column_letter(data.columns.get_loc('Category')+1)}$2:${get_column_letter(data.columns.get_loc('Category')+1)}${n_rows}"
sales_rng = f"{data_range}${sales_col_letter}$2:${sales_col_letter}${n_rows}"
profit_rng = f"{data_range}${profit_col_letter}$2:${profit_col_letter}${n_rows}"
date_col_letter = get_column_letter(date_col)
date_rng = f"{data_range}${date_col_letter}$2:${date_col_letter}${n_rows}"
orderid_rng = f"{data_range}$A$2:$A${n_rows}"

# --- Table A: Revenue & Profit by Region ---
start_row = 4
ws_sum.cell(row=start_row, column=1, value="Revenue & Profit by Region").font = SECTION_FONT
hdr = start_row + 1
headers = ["Region", "Orders", "Total Sales", "Total Profit", "Profit Margin %", "Avg Order Value"]
for i, h in enumerate(headers, start=1):
    ws_sum.cell(row=hdr, column=i, value=h)
style_header_row(ws_sum, hdr, 1, len(headers))
for i, reg in enumerate(regions, start=1):
    r = hdr + i
    ws_sum.cell(row=r, column=1, value=reg)
    ws_sum.cell(row=r, column=2, value=f"=COUNTIF({region_rng},A{r})")
    ws_sum.cell(row=r, column=3, value=f"=SUMIF({region_rng},A{r},{sales_rng})")
    ws_sum.cell(row=r, column=4, value=f"=SUMIF({region_rng},A{r},{profit_rng})")
    ws_sum.cell(row=r, column=5, value=f"=IFERROR(D{r}/C{r},0)")
    ws_sum.cell(row=r, column=6, value=f"=IFERROR(C{r}/B{r},0)")
    ws_sum.cell(row=r, column=3).number_format = "$#,##0"
    ws_sum.cell(row=r, column=4).number_format = "$#,##0"
    ws_sum.cell(row=r, column=5).number_format = "0.0%"
    ws_sum.cell(row=r, column=6).number_format = "$#,##0"
region_table_end = hdr + len(regions)
tot_r = region_table_end + 1
ws_sum.cell(row=tot_r, column=1, value="Total").font = Font(name=FONT_NAME, bold=True)
ws_sum.cell(row=tot_r, column=2, value=f"=SUM(B{hdr+1}:B{region_table_end})")
ws_sum.cell(row=tot_r, column=3, value=f"=SUM(C{hdr+1}:C{region_table_end})")
ws_sum.cell(row=tot_r, column=4, value=f"=SUM(D{hdr+1}:D{region_table_end})")
ws_sum.cell(row=tot_r, column=3).number_format = "$#,##0"
ws_sum.cell(row=tot_r, column=4).number_format = "$#,##0"

# --- Table B: Revenue & Profit by Category ---
start_row2 = tot_r + 3
ws_sum.cell(row=start_row2, column=1, value="Revenue & Profit by Category").font = SECTION_FONT
hdr2 = start_row2 + 1
headers2 = ["Category", "Orders", "Total Sales", "Total Profit", "Profit Margin %"]
for i, h in enumerate(headers2, start=1):
    ws_sum.cell(row=hdr2, column=i, value=h)
style_header_row(ws_sum, hdr2, 1, len(headers2))
for i, cat in enumerate(categories, start=1):
    r = hdr2 + i
    ws_sum.cell(row=r, column=1, value=cat)
    ws_sum.cell(row=r, column=2, value=f"=COUNTIF({category_rng},A{r})")
    ws_sum.cell(row=r, column=3, value=f"=SUMIF({category_rng},A{r},{sales_rng})")
    ws_sum.cell(row=r, column=4, value=f"=SUMIF({category_rng},A{r},{profit_rng})")
    ws_sum.cell(row=r, column=5, value=f"=IFERROR(D{r}/C{r},0)")
    ws_sum.cell(row=r, column=3).number_format = "$#,##0"
    ws_sum.cell(row=r, column=4).number_format = "$#,##0"
    ws_sum.cell(row=r, column=5).number_format = "0.0%"
category_table_end = hdr2 + len(categories)

# --- Table C: Monthly Sales Trend ---
start_row3 = category_table_end + 3
ws_sum.cell(row=start_row3, column=1, value="Monthly Sales Trend").font = SECTION_FONT
hdr3 = start_row3 + 1
ws_sum.cell(row=hdr3, column=1, value="Month")
ws_sum.cell(row=hdr3, column=2, value="Total Sales")
ws_sum.cell(row=hdr3, column=3, value="Total Profit")
style_header_row(ws_sum, hdr3, 1, 3)
for i, m in enumerate(months, start=1):
    r = hdr3 + i
    ws_sum.cell(row=r, column=1, value=m)
    ws_sum.cell(row=r, column=2,
                value=f'=SUMPRODUCT(({date_rng}>=DATE({m[:4]},{m[5:]},1))*'
                      f'({date_rng}<EDATE(DATE({m[:4]},{m[5:]},1),1))*{sales_rng})')
    ws_sum.cell(row=r, column=3,
                value=f'=SUMPRODUCT(({date_rng}>=DATE({m[:4]},{m[5:]},1))*'
                      f'({date_rng}<EDATE(DATE({m[:4]},{m[5:]},1),1))*{profit_rng})')
    ws_sum.cell(row=r, column=2).number_format = "$#,##0"
    ws_sum.cell(row=r, column=3).number_format = "$#,##0"
month_table_end = hdr3 + len(months)

for c, w in zip("ABCDEF", [16, 12, 15, 15, 16, 16]):
    ws_sum.column_dimensions[c].width = w

# ---------------------------------------------------------------------------
# Sheet 3: Lookup Demo (INDEX/MATCH)
# ---------------------------------------------------------------------------
ws_lk = wb.create_sheet("Lookup Demo")
ws_lk["A1"] = "Lookup Functions Demo — INDEX / MATCH"
ws_lk["A1"].font = TITLE_FONT
ws_lk["A2"] = ("Note: XLOOKUP is intentionally not used here because it isn't supported by all "
               "Excel/verification environments (older Excel, some viewers). INDEX/MATCH is the "
               "portable equivalent and works everywhere XLOOKUP does — swap in XLOOKUP(lookup_value, "
               "lookup_array, return_array) directly if your Excel version supports it.")
ws_lk["A2"].font = SUBTITLE_FONT
ws_lk.merge_cells("A2:F2")
ws_lk.row_dimensions[2].height = 30
ws_lk["A2"].alignment = Alignment(wrap_text=True, vertical="top")

# small product reference table (unique products) to look values up FROM
products_unique = data[["Product", "Category"]].drop_duplicates().sort_values("Product").reset_index(drop=True)
avg_price = data.groupby("Product")["UnitPrice"].mean().round(2)
products_unique["AvgUnitPrice"] = products_unique["Product"].map(avg_price)

ref_start = 4
ws_lk.cell(row=ref_start, column=1, value="Product Reference Table").font = SECTION_FONT
hdr_r = ref_start + 1
for i, h in enumerate(["Product", "Category", "AvgUnitPrice"], start=1):
    ws_lk.cell(row=hdr_r, column=i, value=h)
style_header_row(ws_lk, hdr_r, 1, 3)
for i, row in products_unique.iterrows():
    r = hdr_r + 1 + i
    ws_lk.cell(row=r, column=1, value=row["Product"])
    ws_lk.cell(row=r, column=2, value=row["Category"])
    ws_lk.cell(row=r, column=3, value=row["AvgUnitPrice"]).number_format = "$#,##0.00"
ref_end = hdr_r + len(products_unique)

# lookup box: user types a product name in E5, formulas return its category & avg price
ws_lk.cell(row=ref_start, column=5, value="Try it — type a product name below:").font = SECTION_FONT
ws_lk.cell(row=ref_start + 1, column=5, value="Product:")
ws_lk.cell(row=ref_start + 1, column=6, value=products_unique.loc[0, "Product"]).fill = PatternFill("solid", fgColor="FFF2CC")
ws_lk.cell(row=ref_start + 2, column=5, value="Category (INDEX/MATCH):")
ws_lk.cell(row=ref_start + 2, column=6,
           value=f"=INDEX($B${hdr_r+1}:$B${ref_end},MATCH($F${ref_start+1},$A${hdr_r+1}:$A${ref_end},0))")
ws_lk.cell(row=ref_start + 3, column=5, value="Avg Unit Price (INDEX/MATCH):")
ws_lk.cell(row=ref_start + 3, column=6,
           value=f"=INDEX($C${hdr_r+1}:$C${ref_end},MATCH($F${ref_start+1},$A${hdr_r+1}:$A${ref_end},0))")
ws_lk.cell(row=ref_start + 3, column=6).number_format = "$#,##0.00"

for c, w in zip("ABCDEF", [22, 18, 15, 4, 26, 20]):
    ws_lk.column_dimensions[c].width = w

# ---------------------------------------------------------------------------
# Sheet 4: Dashboard — KPI cards + charts
# ---------------------------------------------------------------------------
ws_dash = wb.create_sheet("Dashboard", 0)  # make it the first sheet
ws_dash.sheet_view.showGridLines = False
ws_dash["B2"] = "Sales Performance Dashboard"
ws_dash["B2"].font = Font(name=FONT_NAME, bold=True, size=20, color="1F4E78")
ws_dash["B3"] = "Sample retail dataset · Jan 2024 – Dec 2025"
ws_dash["B3"].font = SUBTITLE_FONT

kpi_defs = [
    ("Total Revenue", f"=SUM(Summary!C{hdr+1}:C{region_table_end})", "$#,##0"),
    ("Total Profit", f"=SUM(Summary!D{hdr+1}:D{region_table_end})", "$#,##0"),
    ("Total Orders", f"=SUM(Summary!B{hdr+1}:B{region_table_end})", "#,##0"),
    ("Overall Margin %", f"=SUM(Summary!D{hdr+1}:D{region_table_end})/SUM(Summary!C{hdr+1}:C{region_table_end})", "0.0%"),
]
kpi_col_start = 2
kpi_row = 5
for i, (label, formula, fmt) in enumerate(kpi_defs):
    col = kpi_col_start + i * 3
    letter = get_column_letter(col)
    letter2 = get_column_letter(col + 1)

    # Fill/border the whole card first
    for rr in range(kpi_row, kpi_row + 4):
        for cc in range(col, col + 2):
            ws_dash.cell(row=rr, column=cc).fill = KPI_FILL
            ws_dash.cell(row=rr, column=cc).border = BORDER

    # Write values into the top-left anchor cells BEFORE merging
    ws_dash.cell(row=kpi_row + 1, column=col, value=label).font = KPI_LABEL_FONT
    val_cell = ws_dash.cell(row=kpi_row + 2, column=col, value=formula)
    val_cell.font = KPI_VALUE_FONT
    val_cell.number_format = fmt

    # Now merge each row of the card across its two columns
    ws_dash.merge_cells(f"{letter}{kpi_row}:{letter2}{kpi_row}")
    ws_dash.merge_cells(f"{letter}{kpi_row+1}:{letter2}{kpi_row+1}")
    ws_dash.merge_cells(f"{letter}{kpi_row+2}:{letter2}{kpi_row+2}")
    ws_dash.merge_cells(f"{letter}{kpi_row+3}:{letter2}{kpi_row+3}")

chart_row = kpi_row + 6

# Bar chart: Sales by Region (from Summary table)
bar = BarChart()
bar.title = "Total Sales by Region"
bar.y_axis.title = "Sales ($)"
bar.style = 10
cats = Reference(ws_sum, min_col=1, min_row=hdr + 1, max_row=region_table_end)
vals = Reference(ws_sum, min_col=3, min_row=hdr, max_row=region_table_end)
bar.add_data(vals, titles_from_data=True)
bar.set_categories(cats)
bar.width, bar.height = 14, 8
ws_dash.add_chart(bar, f"B{chart_row}")

# Line chart: Monthly sales trend
line = LineChart()
line.title = "Monthly Sales Trend"
line.y_axis.title = "Sales ($)"
line.style = 10
cats2 = Reference(ws_sum, min_col=1, min_row=hdr3 + 1, max_row=month_table_end)
vals2 = Reference(ws_sum, min_col=2, min_row=hdr3, max_row=month_table_end)
line.add_data(vals2, titles_from_data=True)
line.set_categories(cats2)
line.width, line.height = 14, 8
ws_dash.add_chart(line, f"H{chart_row}")

# Pie chart: Sales share by Category
pie = PieChart()
pie.title = "Sales Share by Category"
cats3 = Reference(ws_sum, min_col=1, min_row=hdr2 + 1, max_row=category_table_end)
vals3 = Reference(ws_sum, min_col=3, min_row=hdr2, max_row=category_table_end)
pie.add_data(vals3, titles_from_data=True)
pie.set_categories(cats3)
pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
pie.width, pie.height = 14, 8
ws_dash.add_chart(pie, f"B{chart_row + 17}")

ws_dash.column_dimensions["A"].width = 3
for c in "BCDEFGHIJKLMN":
    ws_dash.column_dimensions[c].width = 11

wb.save(OUT_PATH)
print(f"Workbook saved -> {OUT_PATH}")
